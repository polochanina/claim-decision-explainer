"""
Fix mojibake (garbled UTF-8 characters like Ã¥, Ã¤, Ã¶, ÃƒÂ¤) in an Excel file.

Handles both single- and double-encoded corruption, which happens when a
UTF-8 file is misread as Windows-1252 one or more times (common on Mac->Windows
transfers). Only string cells containing the tell-tale 'Ã' marker are touched;
everything else (numbers, dates, formatting, other sheets) is left intact.

Usage:
    python fix_mojibake.py input.xlsx [output.xlsx]

If no output path is given, writes alongside the input as <name>_fixed.xlsx.

Requires:
    pip install ftfy openpyxl
"""

import sys
from pathlib import Path

import ftfy
from openpyxl import load_workbook


def fix_workbook(in_path: str, out_path: str | None = None) -> str:
    in_path = Path(in_path)
    if out_path is None:
        out_path = in_path.with_name(f"{in_path.stem}_fixed{in_path.suffix}")

    wb = load_workbook(in_path)

    fixed = 0
    for ws in wb.worksheets:                       # all sheets
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "Ã" in cell.value:
                    repaired = ftfy.fix_text(cell.value)
                    if repaired != cell.value:
                        cell.value = repaired
                        fixed += 1

    wb.save(out_path)

    # verify nothing was missed
    check = load_workbook(out_path)
    remaining = sum(
        1
        for ws in check.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and "Ã" in cell.value
    )

    print(f"Repaired {fixed} cell(s). Remaining mojibake cells: {remaining}")
    print(f"Saved: {out_path}")
    return str(out_path)


if __name__ == "__main__":
    input_path = "claim_use_case_dataset.xlsx"
    output_path = "dataset.xlsx"
    fix_workbook(input_path, output_path)